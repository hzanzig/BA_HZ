import csv
import mimetypes
import string
from openpyxl import load_workbook
import xlrd
import pandas

ICDCatalog_object = load_workbook(r'C:\Users\handb\Documents\Bachelorarbeit\ICDandOPS\ICD-10-GM.xlsx')
ICDCharite_object = load_workbook(r'C:\Users\handb\Documents\Bachelorarbeit\ICDandOPS\icd_diagnosis_full_charite.xlsx')

ICDCharite_worksheet = ICDCharite_object['icd_diagnosis_full_charite']
ICDCatalog_worksheet = ICDCatalog_object['Tabelle1']

ICDCatalog_Zeilenzahl = ICDCatalog_worksheet.max_row
ICDCharite_Zeilenzahl = ICDCharite_worksheet.max_row
matches = 0
mismatches = 0
listofmismatches=[]

for j in range(1,ICDCatalog_Zeilenzahl+1):
    x = matches
    y = mismatches
    for i in range(1,ICDCharite_Zeilenzahl+1):
        if (ICDCharite_worksheet[('A'+str(i))].value)==(ICDCatalog_worksheet[('A'+str(j))].value):
            matches = matches+1
            break
    if matches == x:
        mismatches +=1
        listofmismatches.append(ICDCatalog_worksheet[('A'+str(j))].value)
    print(j,matches,mismatches)

print(matches)
print(mismatches)









#for i in range(1,ICDCatalog_Zeilenzahl-1):
#    x = mismatches
#    y = matches
#    for j in range(1,ICDCharite_Zeilenzahl-1):
#        if (ICDCharite_worksheet[('A'+str(j))].value)==(ICDCatalog_worksheet[('A'+str(i))].value):
#            matches = matches+1
#            break
#        elif j == ICDCharite_Zeilenzahl-1:
#            mismatches +=1
#            listofmismatches.append(ICDCatalog_worksheet[('A'+str(i))].value)
#            break
#    print(y,x,j)

#for i in range(1,ICDCatalog_Zeilenzahl-1):
#    for j in range(1,ICDCharite_Zeilenzahl-1):
#        if (ICDCharite_worksheet[('A'+str(j))].value)==(ICDCatalog_worksheet[('A'+str(ji))].value):
 #           matches = matches+1
 #           break
 #       elif j == ICDCharite_Zeilenzahl-1:
 #           mismatches +=1
 #           listofmismatches.append(ICDCatalog_worksheet[('A'+str(i))].value)
 #           break
 #   print(y,x,j)


#    x = mismatches
#    y = matches
#    j = 1
#    while (x<mismatches or y<matches):# and j<=ICDCharite_Zeilenzahl-1
#        if (ICDCharite_worksheet[('A'+str(i))].value)==(ICDCatalog_worksheet[('A'+str(j))].value):
#            matches = matches+1
#            print(y,matches)
#        elif j == ICDCharite_Zeilenzahl-1:
#            mismatches +=1
#            print(x,mismatches)
#           listofmismatches.append(ICDCatalog_worksheet[('A'+str(j))].value)
#        j+=1
#    print(y,x,j)

#print(ICDCharite_worksheet[('A'+str(ICDCharite_worksheet))].value)

#if (ICDCharite_worksheet['A500'].value)==(ICDCatalog_worksheet['A484'].value):
#    print("WORKS")
#else:
#   print(ICDCharite_worksheet['A500'].value)
#   print(ICDCatalog_worksheet['A484'].value)
#for i in ICDCharite_worksheet.values:
#    print(i)


