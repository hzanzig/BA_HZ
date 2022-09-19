#!/usr/bin/python
import psycopg2
import xlwt
from openpyxl import load_workbook
from xlwt import Workbook
from config import config

#loading of result excel file
OPSCharite_object = load_workbook(r'C:\Users\handb\Documents\GitHub\BA_HZ\OPSCharite_result.xlsx')
OPSCharite_result_worksheet1 = OPSCharite_object['Sheet 1']
OPSCharite_result_worksheet2 = OPSCharite_object['Sheet 2']
OPSCharite_result_worksheet3 = OPSCharite_object['Sheet 3']
OPSCharite_result_worksheet4 = OPSCharite_object['Sheet 4']
worksheets = [OPSCharite_result_worksheet1,OPSCharite_result_worksheet2,OPSCharite_result_worksheet3,OPSCharite_result_worksheet4]

#loading of second iteration result excel file
OPSCharite_object2 = load_workbook(r'C:\Users\handb\Documents\GitHub\BA_HZ\OPSCharite_result_2.xlsx')
OPSCharite_result2_worksheet1 = OPSCharite_object2['Sheet 1']
OPSCharite_result2_worksheet2 = OPSCharite_object2['Sheet 2']
OPSCharite_result2_worksheet3 = OPSCharite_object2['Sheet 3']
OPSCharite_result2_worksheet4 = OPSCharite_object2['Sheet 4']
worksheets2 = [OPSCharite_result2_worksheet1,OPSCharite_result2_worksheet2,OPSCharite_result2_worksheet3,OPSCharite_result2_worksheet4]

#loading of third iteration result excel file
OPSCharite_object3 = load_workbook(r'C:\Users\handb\Documents\GitHub\BA_HZ\OPSCharite_result_3.xlsx')
OPSCharite_result3_worksheet1 = OPSCharite_object3['Sheet 1']
OPSCharite_result3_worksheet2 = OPSCharite_object3['Sheet 2']
OPSCharite_result3_worksheet3 = OPSCharite_object3['Sheet 3']
OPSCharite_result3_worksheet4 = OPSCharite_object3['Sheet 4']
worksheets3 = [OPSCharite_result3_worksheet1,OPSCharite_result3_worksheet2,OPSCharite_result3_worksheet3,OPSCharite_result3_worksheet4]

matches = 0
mismatches = 0
mismatchlist = []

matches_2it = 0
mismatches_2it = 0
mismatchlist_2it = []

matches_3it = 0
mismatches_3it = 0
mismatchlist_3it = []

for i in range(0,4):
    for j in range(2,worksheets[i].max_row):
        singleworksheet=worksheets[i]
        if singleworksheet[('C'+str(j))].value == 1:
            matches +=1
        else:
            mismatches +=1
            mismatchlist.append(singleworksheet[('A'+str(j))].value)
        singleworksheet2_it=worksheets2[i]
        if singleworksheet2_it[('E'+str(j))].value == 0:
            mismatches_2it +=1
            mismatchlist_2it.append(singleworksheet2_it[('A'+str(j))].value)
        else:
            matches_2it +=1
        singleworksheet3_it=worksheets3[i]
        if singleworksheet3_it[('E'+str(j))].value == 0:
            mismatches_3it +=1
            mismatchlist_3it.append(singleworksheet3_it[('A'+str(j))].value)
        else:
            matches_3it +=1

print("Matches: ", matches)
print("Mismatches: ", mismatches)
#print(mismatchlist)

print("Matches2: ", matches_2it)
print("Mismatches2: ", mismatches_2it)
#print(mismatchlist_2it)

print("Matches3: ", matches_3it)
print("Mismatches3: ", mismatches_3it)
#print(mismatchlist_3it)

#1
#Matches:  121649
#Mismatches:  89610

#result2
#Matches2:  199128
#Mismatches2:  12131

#result3
#Matches3:  204888
#Mismatches3:  6371