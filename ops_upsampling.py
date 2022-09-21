#!/usr/bin/python
from numpy import integer
from unittest import skipIf
import psycopg2
import xlwt
from openpyxl import load_workbook
from xlwt import Workbook
from config import config
import pandas as pd
from operator import concat

# loading the result excel file of the mapping process
OPSCharite_object3 = load_workbook(r'C:\Users\handb\Documents\GitHub\BA_HZ\OPSCharite_result_3.xlsx')
OPSCharite_result3_worksheet1 = OPSCharite_object3['Sheet 1']
OPSCharite_result3_worksheet2 = OPSCharite_object3['Sheet 2']
OPSCharite_result3_worksheet3 = OPSCharite_object3['Sheet 3']
OPSCharite_result3_worksheet4 = OPSCharite_object3['Sheet 4']
worksheets3 = [OPSCharite_result3_worksheet1,OPSCharite_result3_worksheet2,OPSCharite_result3_worksheet3,OPSCharite_result3_worksheet4]
#creating the result dataframe of the upsampling process
ResultsDataframeops = pd.DataFrame({"c_procedure_1":[], "c_procedure_katalog_1":[],"mapped_katalog":[],"number_of_mappings":[],"number_of_upsamplings":[],"matchable_code":[]})

conn = None
params = config()
# connect to the PostgreSQL server
print('Connecting to the PostgreSQL database...')
conn = psycopg2.connect(**params)
# create a cursor
cur = conn.cursor()


for i in range(0,len(worksheets3)):
    for j in range(2,worksheets3[i].max_row):
        excelsheet=worksheets3[i]
        ops_value = str(excelsheet[('A'+str(j))].value)
        try:
            if str(excelsheet[('E'+str(j))].value) == '0':
                # check if the mapping was unsuccessful
                print("sheet"+str(i+1)+","+str(j)+"/4,"+str(worksheets3[i].max_row))
                ops_valuestrip=ops_value
                upsampling = 0
                whileexit = 0
                # shortening the ops code with every while iteration
                while whileexit ==0:
                    ops_valuestrip=ops_valuestrip[:-1]
                    query = "".join(['SELECT'," * FROM public.concept WHERE vocabulary_id LIKE 'OPS%' AND concept_code LIKE '", str(ops_valuestrip), "%'"])
                    #the sql query searches with a wildcard at the end of the stripped code in the database
                    cur.execute(query)
                    QueryErgebnis = cur.fetchall()
                    upsampling = upsampling+1
                    whileexit = cur.rowcount
                opsKatalog =[]
                #as many results can occur, which could corrupt the excel table format, the found catalogs will be reduced to the first 100.
                if cur.rowcount > 100:
                    for z in range(0,100):
                        opsKatalog.append(QueryErgebnis[z][3])
                else:
                    for z in range(0,cur.rowcount):
                        opsKatalog.append(QueryErgebnis[z][3])
                toappend = [ops_value, excelsheet[('B'+str(j))].value, str(opsKatalog),cur.rowcount,upsampling,str(ops_valuestrip)]
                ResultsDataframeops = ResultsDataframeops.append(pd.Series(toappend, index=ResultsDataframeops.columns[:len(toappend)]), ignore_index=True)
            else:
                skipIf
        except:
            toappend = [ops_value, excelsheet[('B'+str(j))].value,"NA","NA","NA","NA"]
            ResultsDataframeops = ResultsDataframeops.append(pd.Series(toappend, index=ResultsDataframeops.columns[:len(toappend)]), ignore_index=True)

ResultsDataframeops.to_csv('OPS_upsampling.csv')
# close the communication with the PostgreSQL
cur.close()
if conn is not None:
    conn.close()

