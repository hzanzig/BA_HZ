#!/usr/bin/python
import psycopg2
import xlwt
from openpyxl import load_workbook
from config import config
import pandas as pd

#load of Charité Workbook and creation of results-dataframe
OPSCharite_object = load_workbook(r'C:\Users\handb\Documents\GitHub\BA_HZ\ops_codes_full_charite.xlsx')
OPSCharite_worksheet = OPSCharite_object['ops_codes_full_charite']
OPSCharite_Zeilenzahl = OPSCharite_worksheet.max_row
ResultsDataframeOPS = pd.DataFrame({"c_procedure_1":[], "c_procedure_katalog_1":[], "mapped_katalog":[],"number_of_mappings":[]})

# connect to the PostgreSQL server
conn = None
params = config()
conn = psycopg2.connect(**params)
cur = conn.cursor()

for i in range(2,OPSCharite_Zeilenzahl):
    print(str(i)+"/"+str(OPSCharite_Zeilenzahl))
    current_ops_value=OPSCharite_worksheet[('A'+str(i))].value
    current_ops_value = current_ops_value.lower() #all letters to lowercase, as the SQL querry is case sensitive and the codes from ATHENA only consist of lowercase 
    query = "".join(['SELECT'," * FROM public.concept WHERE vocabulary_id LIKE 'OPS%' AND concept_code LIKE '", str(current_ops_value), "%'"])
    cur.execute(query)
    QueryErgebnis = cur.fetchall()
    OPSKatalog =[]
    for z in range(0,cur.rowcount):
        OPSCatalog.append(QueryErgebnis[z][3])
    toappend = [current_ops_value, OPSCharite_worksheet[('B'+str(i))].value, str(OPSCatalog),cur.rowcount]
    ResultsDataframeOPS = ResultsDataframeOPS.append(pd.Series(toappend, index=ResultsDataframeOPS.columns[:len(toappend)]), ignore_index=True)

ResultsDataframeOPS.to_csv('OPS_full_results.csv')

# close the communication with the PostgreSQL
cur.close()
if conn is not None:
    conn.close()
