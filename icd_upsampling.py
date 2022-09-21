#!/usr/bin/python
from numpy import integer
from unittest import skipIf
import psycopg2
import xlwt
from openpyxl import load_workbook
from xlwt import Workbook
from config import config
import pandas as pd
from operator import concat

# loading the result excel file of the mapping process
ICDCharite_object = load_workbook(r'C:\Users\handb\Documents\GitHub\BA_HZ\ICDCharite_result.xlsx')
ICDCharite_result_worksheet1 = ICDCharite_object['Sheet 1']
ICDCharite_result_worksheet2 = ICDCharite_object['Sheet 2']
ICDCharite_result_worksheet3 = ICDCharite_object['Sheet 3']
ICDCharite_result_worksheet4 = ICDCharite_object['Sheet 4']
worksheets = [ICDCharite_result_worksheet1,ICDCharite_result_worksheet2,ICDCharite_result_worksheet3,ICDCharite_result_worksheet4]
#creating the result dataframe of the upsampling process
ResultsDataframeICD = pd.DataFrame({"c_diagnose_1":[], "c_diagnose_Catalog_1":[],"mapped_Catalog":[],"number_of_mappings":[],"number_of_upsamplings":[],"matchable_code":[]})

conn = None
params = config()
# connect to the PostgreSQL server
print('Connecting to the PostgreSQL database...')
conn = psycopg2.connect(**params)
# create a cursor
cur = conn.cursor()

for i in range(0,len(worksheets)):
    for j in range(2,worksheets[i].max_row):
        excelsheet=worksheets[i]
        icd_value = str(excelsheet[('A'+str(j))].value)
        try:
            # check if the mapping was unsuccessful
            if str(excelsheet[('C'+str(j))].value) == '0':
                icd_valuestrip=icd_value
                upsampling = 0
                whileexit = 0
                # shortening the icd code with every while iteration
                while whileexit ==0:
                    icd_valuestrip=icd_valuestrip[:-1]
                    #the sql query searches with a wildcard at the end of the stripped code in the database
                    query = "".join(['SELECT'," * FROM public.concept WHERE vocabulary_id LIKE 'ICD%' AND concept_code LIKE '", str(icd_valuestrip), "%'"])
                    cur.execute(query)
                    QueryErgebnis = cur.fetchall()
                    upsampling = upsampling+1
                    whileexit = cur.rowcount
                ICDCatalog =[]
                #as many results can occur, which could corrupt the excel table format, the found catalogs will be reduced to the first 100.
                if cur.rowcount > 100:
                    for z in range(0,100):
                        ICDCatalog.append(QueryErgebnis[z][3])
                else:
                    for z in range(0,cur.rowcount):
                        ICDCatalog.append(QueryErgebnis[z][3])
                toappend = [icd_value, excelsheet[('B'+str(j))].value,ICDCatalog, cur.rowcount,upsampling,str(icd_valuestrip)]
                ResultsDataframeICD = ResultsDataframeICD.append(pd.Series(toappend, index=ResultsDataframeICD.columns[:len(toappend)]), ignore_index=True)
            else:
                skipIf
        except:
            toappend = [icd_value, excelsheet[('B'+str(j))].value,"NA","NA","NA","NA"]
            ResultsDataframeICD = ResultsDataframeICD.append(pd.Series(toappend, index=ResultsDataframeICD.columns[:len(toappend)]), ignore_index=True)

ResultsDataframeICD.to_csv('ICD_upsampling.csv')
# close the communication with the PostgreSQL
cur.close()
if conn is not None:
    conn.close()
