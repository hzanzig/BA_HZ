#!/usr/bin/python
from numpy import integer
import psycopg2
import xlwt
from openpyxl import load_workbook
from xlwt import Workbook
from config import config

OPSCharite_object = load_workbook(r'C:\GitHub\BA_HZ\OPSCharite_result_2.xlsx')
#creation of new result file
OPSCharite_result_3 = Workbook()
sheet1 = OPSCharite_result_3.add_sheet('Sheet 1')
sheet2 = OPSCharite_result_3.add_sheet('Sheet 2')
sheet3 = OPSCharite_result_3.add_sheet('Sheet 3')
sheet4 = OPSCharite_result_3.add_sheet('Sheet 4')
sheets = [sheet1,sheet2,sheet3,sheet4]

#loading the excel sheets of the second run
OPSCharite_result_worksheet = OPSCharite_object['Sheet 1']
OPSCharite_result_worksheet2 = OPSCharite_object['Sheet 2']
OPSCharite_result_worksheet3 = OPSCharite_object['Sheet 3']
OPSCharite_result_worksheet4 = OPSCharite_object['Sheet 4']
OPSCharite_Rowcount1 = OPSCharite_result_worksheet.max_row
OPSCharite_Rowcount2 = OPSCharite_result_worksheet2.max_row
OPSCharite_Rowcount3 = OPSCharite_result_worksheet3.max_row
OPSCharite_Rowcount4 = OPSCharite_result_worksheet4.max_row

#connection to database
conn = None
params = config()
# connect to the PostgreSQL server
print('Connecting to the PostgreSQL database...')
conn = psycopg2.connect(**params)
# create a cursor
cur = conn.cursor()
        
#set-up of the result sheets 
for i in range(0,4):
    sheets[i].write(0,0,"c_procedure_1")
    sheets[i].write(0,1,"c_procedure_catalog_1")
    sheets[i].write(0,2,"mapping_success")
    sheets[i].write(0,3,"mapped_catalog")
    sheets[i].write(0,4,"number_of_mappings")

#checking the first sheet of the second run
for i in range(2,OPSCharite_Rowcount1):
    print("Blatt1 "+str(i)+"/"+str(OPSCharite_Rowcount1))
    ops_value=OPSCharite_result_worksheet[('A'+str(i))].value
    #checking if OPS code was already a match or not
    if OPSCharite_result_worksheet[('E'+str(i))].value==1:
        sheet1.write(i-1, 0, ops_value)
        sheet1.write(i-1, 1, OPSCharite_result_worksheet[('B'+str(i))].value)
        sheet1.write(i-1, 2, OPSCharite_result_worksheet[('C'+str(i))].value)
        sheet1.write(i-1, 3, OPSCharite_result_worksheet[('D'+str(i))].value)
        sheet1.write(i-1, 4, OPSCharite_result_worksheet[('E'+str(i))].value)
    else:
        ops_value = ops_value.lower() #making entire string lowercase as SQL query is casesensitive and the letters are lowercase in ATHENA vocabulary
        query = "".join(['SELECT'," * FROM public.concept WHERE vocabulary_id LIKE 'OPS%' AND concept_code LIKE '", str(ops_value), "%'"])
        cur.execute(query)
        QueryResult = cur.fetchall()
        if cur.rowcount == 0:
            sheet1.write(i-1, 0,  ops_value)
            sheet1.write(i-1, 1,  OPSCharite_result_worksheet[('B'+str(i))].value)
            sheet1.write(i-1, 2, 0)
        else:
            sheet1.write(i-1, 0,  ops_value)
            sheet1.write(i-1, 1,  OPSCharite_result_worksheet[('B'+str(i))].value)
            sheet1.write(i-1, 2, 1)
            OPSCatalog =[]
            for z in range(0,cur.rowcount):
                OPSCatalog.append(QueryResult[z][3])
            sheet1.write(i-1, 3, str(OPSCatalog))
            sheet1.write(i-1,4,cur.rowcount)

#checking the second sheet of the second run
for i in range(2,OPSCharite_Rowcount2):
    print("Blatt2 "+str(i)+"/"+str(OPSCharite_Rowcount2))
    ops_value=OPSCharite_result_worksheet2[('A'+str(i))].value
    #checking if OPS code was already a match or not
    if OPSCharite_result_worksheet2[('E'+str(i))].value==1:
        sheet2.write(i-1, 0, ops_value)
        sheet2.write(i-1, 1, OPSCharite_result_worksheet2[('B'+str(i))].value)
        sheet2.write(i-1, 2, OPSCharite_result_worksheet2[('C'+str(i))].value)
        sheet2.write(i-1, 3, OPSCharite_result_worksheet2[('D'+str(i))].value)
        sheet2.write(i-1, 4, OPSCharite_result_worksheet2[('E'+str(i))].value)
    else:
        ops_value = ops_value.lower() #making entire string lowercase as SQL query is casesensitive and the letters are lowercase in ATHENA vocabulary
        query = "".join(['SELECT'," * FROM public.concept WHERE vocabulary_id LIKE 'OPS%' AND concept_code LIKE '", str(ops_value), "'"])
        cur.execute(query)
        QueryResult = cur.fetchall()
        if cur.rowcount == 0:
            sheet2.write(i-1, 0,  ops_value)
            sheet2.write(i-1, 1,  OPSCharite_result_worksheet2[('B'+str(i))].value)
            sheet2.write(i-1, 2, 0)
        else:
            sheet2.write(i-1, 0,  ops_value)
            sheet2.write(i-1, 1,  OPSCharite_result_worksheet2[('B'+str(i))].value)
            sheet2.write(i-1, 2, 1)
            OPSCatalog =[]
            for z in range(0,cur.rowcount):
                OPSCatalog.append(QueryResult[z][3])
            sheet2.write(i-1, 3, str(OPSCatalog))
            sheet2.write(i-1,4,cur.rowcount)

#checking the third sheet of the second run
for i in range(2,OPSCharite_Rowcount3):
    print("Blatt3 "+str(i)+"/"+str(OPSCharite_Rowcount3))
    ops_value=OPSCharite_result_worksheet3[('A'+str(i))].value
    #checking if OPS code was already a match or not
    if OPSCharite_result_worksheet3[('E'+str(i))].value==1:
            sheet3.write(i-1, 0, ops_value)
            sheet3.write(i-1, 1, OPSCharite_result_worksheet3[('B'+str(i))].value)
            sheet3.write(i-1, 2, OPSCharite_result_worksheet3[('C'+str(i))].value)
            sheet3.write(i-1, 3, OPSCharite_result_worksheet3[('D'+str(i))].value)
            sheet3.write(i-1, 4, OPSCharite_result_worksheet3[('E'+str(i))].value)
    else:
        ops_value = ops_value.lower() #making entire string lowercase as SQL query is casesensitive and the letters are lowercase in ATHENA vocabulary
        query = "".join(['SELECT'," * FROM public.concept WHERE vocabulary_id LIKE 'OPS%' AND concept_code LIKE '", str(ops_value), "'"])
        cur.execute(query)
        QueryResult = cur.fetchall()
        if cur.rowcount == 0:
            sheet3.write(i-1, 0,  ops_value)
            sheet3.write(i-1, 1,  OPSCharite_result_worksheet3[('B'+str(i))].value)
            sheet3.write(i-1, 2, 0)
        else:
            sheet3.write(i-1, 0,  ops_value)
            sheet3.write(i-1, 1,  OPSCharite_result_worksheet3[('B'+str(i))].value)
            sheet3.write(i-1, 2, 1)
            OPSCatalog =[]
            for z in range(0,cur.rowcount):
                OPSCatalog.append(QueryResult[z][3])
            sheet3.write(i-1, 3, str(OPSCatalog))
            sheet3.write(i-1,4,cur.rowcount)

#checking the fourth sheet of the second run
for i in range(2,OPSCharite_Rowcount4):
    print("Blatt4 "+str(i)+"/"+str(OPSCharite_Rowcount4))
    ops_value=OPSCharite_result_worksheet4[('A'+str(i))].value
    #checking if OPS code was already a match or not
    if OPSCharite_result_worksheet4[('E'+str(i))].value==1:
            sheet4.write(i-1, 0, ops_value)
            sheet4.write(i-1, 1, OPSCharite_result_worksheet4[('B'+str(i))].value)
            sheet4.write(i-1, 2, OPSCharite_result_worksheet4[('C'+str(i))].value)
            sheet4.write(i-1, 3, OPSCharite_result_worksheet4[('D'+str(i))].value)
            sheet4.write(i-1, 4, OPSCharite_result_worksheet4[('E'+str(i))].value)
    else:
        ops_value = ops_value.lower() #making entire string lowercase as SQL query is casesensitive and the letters are lowercase in ATHENA vocabulary
        query = "".join(['SELECT'," * FROM public.concept WHERE vocabulary_id LIKE 'OPS%' AND concept_code LIKE '", str(ops_value), "'"])
        cur.execute(query)
        QueryResult = cur.fetchall()
        if cur.rowcount == 0:
            sheet4.write(i-1, 0,  ops_value)
            sheet4.write(i-1, 1,  OPSCharite_result_worksheet4[('B'+str(i))].value)
            sheet4.write(i-1, 2, 0)
        else:
            sheet4.write(i-1, 0,  ops_value)
            sheet4.write(i-1, 1,  OPSCharite_result_worksheet4[('B'+str(i))].value)
            sheet4.write(i-1, 2, 1)
            OPSCatalog =[]
            for z in range(0,cur.rowcount):
                OPSCatalog.append(QueryResult[z][3])
            sheet4.write(i-1, 3, str(OPSCatalog))
            sheet4.write(i-1,4,cur.rowcount)


OPSCharite_result_3.save('OPSCharite_result_3.xls')

# close the communication with the PostgreSQL
cur.close()
if conn is not None:
    conn.close()
    print('Database connection closed.')

