#!/usr/bin/python
import psycopg2
import xlwt
from openpyxl import load_workbook
from xlwt import Workbook
from config import config
import pandas as pd

#loading the OPS and ICD result Excel files, which are made of 4 differfent sheets, each filled with data
ICDCharite_object = load_workbook(r'C:\GitHub\BA_HZ\ICDCharite_result.xlsx')
ICDCharite_result_worksheet1 = ICDCharite_object['Sheet 1']
ICDCharite_result_worksheet2 = ICDCharite_object['Sheet 2']
ICDCharite_result_worksheet3 = ICDCharite_object['Sheet 3']
ICDCharite_result_worksheet4 = ICDCharite_object['Sheet 4']
worksheets = [ICDCharite_result_worksheet1,ICDCharite_result_worksheet2,ICDCharite_result_worksheet3,ICDCharite_result_worksheet4]
OPSCharite_object3 = load_workbook(r'C:\GitHub\BA_HZ\OPSCharite_result_3.xlsx')
OPSCharite_result3_worksheet1 = OPSCharite_object3['Sheet 1']
OPSCharite_result3_worksheet2 = OPSCharite_object3['Sheet 2']
OPSCharite_result3_worksheet3 = OPSCharite_object3['Sheet 3']
OPSCharite_result3_worksheet4 = OPSCharite_object3['Sheet 4']
worksheets3 = [OPSCharite_result3_worksheet1,OPSCharite_result3_worksheet2,OPSCharite_result3_worksheet3,OPSCharite_result3_worksheet4]

#creation of data frame for ICD and OPS 
dataframeICD = pd.DataFrame({ICDCharite_result_worksheet1[('A1')].value:[], ICDCharite_result_worksheet1[('B1')].value:[], ICDCharite_result_worksheet1[('C1')].value:[],ICDCharite_result_worksheet1[('D1')].value:[],ICDCharite_result_worksheet1[('E1')].value:[]})
dataframeOPS = pd.DataFrame({OPSCharite_result3_worksheet1[('A1')].value:[], OPSCharite_result3_worksheet1[('B1')].value:[], OPSCharite_result3_worksheet1[('C1')].value:[],OPSCharite_result3_worksheet1[('D1')].value:[],OPSCharite_result3_worksheet1[('E1')].value:[]})

#filling the ops data frame with the input
for i in range(0,4):
    for j in range(2,worksheets3[i].max_row):
        test=worksheets3[i]
        toappend = [test[('A'+str(j))].value, test[('B'+str(j))].value, test[('C'+str(j))].value,test[('D'+str(j))].value,test[('E'+str(j))].value]
        dataframeOPS = dataframeOPS.append(pd.Series(toappend, index=dataframeOPS.columns[:len(toappend)]), ignore_index=True)

dataframeOPS.to_csv('ops_full_result.csv')


#filling the icd data frame with the input
for i in range(0,4):
    for j in range(2,worksheets[i].max_row):
        test=worksheets[i]
        toappend = [test[('A'+str(j))].value, test[('B'+str(j))].value, test[('C'+str(j))].value,test[('D'+str(j))].value,test[('E'+str(j))].value]
        dataframeICD = dataframeICD.append(pd.Series(toappend, index=dataframeICD.columns[:len(toappend)]), ignore_index=True)
        #dataframeICD.loc[dataframeICD.shape[0]] = [test[('A'+str(j))].value, test[('B'+str(j))].value, test[('C'+str(j))].value,test[('D'+str(j))].value,test[('E'+str(j))].value]

dataframeICD.to_csv('icd_full_result.csv')
