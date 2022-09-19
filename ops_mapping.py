#!/usr/bin/python
import psycopg2
import xlwt
from openpyxl import load_workbook
from xlwt import Workbook
from config import config

OPSCharite_object = load_workbook(r'C:\GitHub\BA_HZ\ops_codes_full_charite.xlsx')
OPSCharite_result = Workbook()
sheet1 = OPSCharite_result.add_sheet('Sheet 1')
sheet2 = OPSCharite_result.add_sheet('Sheet 2')
sheet3 = OPSCharite_result.add_sheet('Sheet 3')
sheet4 = OPSCharite_result.add_sheet('Sheet 4')
sheets = [sheet1,sheet2,sheet3,sheet4]
OPSCharite_worksheet = OPSCharite_object['ops_codes_full_charite']
OPSCharite_Rowcount = OPSCharite_worksheet.max_row

#connection to database
conn = None
params = config()
# connect to the PostgreSQL server
print('Connecting to the PostgreSQL database...')
conn = psycopg2.connect(**params)
# create a cursor
cur = conn.cursor()
        
#set-up of the result sheets 
for i in range(0,4):
    sheets[i].write(0,0,"c_procedure_1")
    sheets[i].write(0,1,"c_procedure_catalog_1")
    sheets[i].write(0,2,"mapping_success")
    sheets[i].write(0,3,"mapped_catalog")
    sheets[i].write(0,4,"number_of_mappings")

#begin of the iteraton
for i in range(2,OPSCharite_Rowcount):
    print(str(i)+"/"+str(OPSCharite_Rowcount)) #print of iterations 
    ops_value=OPSCharite_worksheet[('A'+str(i))].value #getting the OPS code
    query = "".join(['SELECT'," * FROM public.concept WHERE vocabulary_id LIKE 'OPS%' AND concept_code LIKE '", str(ops_value), "'"])
    cur.execute(query)  #executing the SELECT query
    QueryResult = cur.fetchall()
    #Workbook can't extend the rowcount of 65,536. For number of OPS codes, several EXCEL sheets were therefore neccessary
    if i<65500:
        if cur.rowcount == 0:
            sheet1.write(i-1, 0,  ops_value)
            sheet1.write(i-1, 1,  OPSCharite_worksheet[('B'+str(i))].value)
            sheet1.write(i-1, 2, 0)
        else:
            sheet1.write(i-1, 0,  ops_value)
            sheet1.write(i-1, 1,  OPSCharite_worksheet[('B'+str(i))].value)
            sheet1.write(i-1, 2, 1)
            OPSCatalog =[]
            for z in range(0,cur.rowcount):
                OPSCatalog.append(QueryResult[z][3])
            sheet1.write(i-1, 3, str(OPSCatalog))
            sheet1.write(i-1,4,cur.rowcount)
    elif i<131000:
        if cur.rowcount == 0:
            sheet2.write(i-65499, 0,  ops_value)
            sheet2.write(i-65499, 1,  OPSCharite_worksheet[('B'+str(i))].value)
            sheet2.write(i-65499, 2, 0)
        else:
            sheet2.write(i-65499, 0,  ops_value)
            sheet2.write(i-65499, 1,  OPSCharite_worksheet[('B'+str(i))].value)
            sheet2.write(i-65499, 2, 1)
            OPSCatalog =[]
            for z in range(0,cur.rowcount):
                OPSCatalog.append(QueryResult[z][3])
            sheet2.write(i-65499, 3, str(OPSCatalog))
            sheet2.write(i-65499,4,cur.rowcount)
    elif i<196000:
        if cur.rowcount == 0:
            sheet3.write(i-130999, 0,  ops_value)
            sheet3.write(i-130999, 1,  OPSCharite_worksheet[('B'+str(i))].value)
            sheet3.write(i-130999, 2, 0)
        else:
            sheet3.write(i-130999, 0,  ops_value)
            sheet3.write(i-130999, 1,  OPSCharite_worksheet[('B'+str(i))].value)
            sheet3.write(i-130999, 2, 1)
            OPSCatalog =[]
            for z in range(0,cur.rowcount):
                OPSCatalog.append(QueryResult[z][3])
            sheet3.write(i-130999, 3, str(OPSCatalog))
            sheet3.write(i-130999,4,cur.rowcount)
    else:
        if cur.rowcount == 0:
            sheet4.write(i-195999, 0,  ops_value)
            sheet4.write(i-195999, 1,  OPSCharite_worksheet[('B'+str(i))].value)
            sheet4.write(i-195999, 2, 0)
        else:
            sheet4.write(i-195999, 0,  ops_value)
            sheet4.write(i-195999, 1,  OPSCharite_worksheet[('B'+str(i))].value)
            sheet4.write(i-195999, 2, 1)
            OPSCatalog =[]
            for z in range(0,cur.rowcount):
                OPSCatalog.append(QueryResult[z][3])
            sheet4.write(i-195999, 3, str(OPSCatalog))
            sheet4.write(i-195999,4,cur.rowcount)


OPSCharite_result.save('OPSCharite_result.xls')

#close the communication with the PostgreSQL
cur.close()
if conn is not None:
    conn.close()
    print('Database connection closed.')

