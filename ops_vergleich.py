import csv
import mimetypes
import string
from openpyxl import load_workbook
import xlrd
import xlwt
from xlwt import Workbook
import pandas
import pandas as pd

OPSCatalog_object = load_workbook(r'C:\Users\handb\Documents\Bachelorarbeit\ICDandOPS\opskatalog.xlsx')
OPSCharite_object = load_workbook(r'C:\Users\handb\Documents\Bachelorarbeit\ICDandOPS\ops_codes_full_charite.xlsx')
#riter = pandas.ExcelWriter(r'C:\Users\handb\Documents\Bachelorarbeit\ICDandOPS\ops_codes_full_charite.xlsx', engine='openpyxl')

OPSCharite_worksheet = OPSCharite_object['ops_codes_full_charite']
OPSCatalog_worksheet = OPSCatalog_object['Sheet1']

wb = Workbook()

sheet1 = wb.add_sheet('Sheet1')


OPSCatalog_Zeilenzahl = OPSCatalog_worksheet.max_row
OPSCharite_Zeilenzahl = OPSCharite_worksheet.max_row
matches = 0
mismatches = 0
listofmismatches=[]

for i in range(1,OPSCharite_Zeilenzahl+1):
    x = matches
    for j in range(1,OPSCatalog_Zeilenzahl+1):
        if (OPSCharite_worksheet[('A'+str(i))].value)==(OPSCatalog_worksheet[('A'+str(j))].value):
            matches = matches+1
            sheet1.write(i-1, 0, (OPSCharite_worksheet[('A'+str(i))].value))
            sheet1.write(i-1, 1, 1)
            break
    if matches == x:
        mismatches +=1
        sheet1.write(i-1, 0, (OPSCharite_worksheet[('A'+str(i))].value))
        sheet1.write(i-1, 1, 0)
    print(j,matches,mismatches)

wb.save('opsDataMapTest.xls')
print(matches)
print(mismatches)
