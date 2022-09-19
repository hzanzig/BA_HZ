#!/usr/bin/python
import psycopg2
import xlwt
from openpyxl import load_workbook
from config import config
import pandas as pd

#load of Charité Workbook and creation of results-dataframe
ICDCharite_object = load_workbook(r'C:\Users\handb\Documents\GitHub\BA_HZ\icd_diagnosis_full_charite.xlsx')
ICDCharite_worksheet = ICDCharite_object['icd_diagnosis_full_charite']
ICDCharite_Zeilenzahl = ICDCharite_worksheet.max_row
ResultsDataframeICD = pd.DataFrame({"c_diagnose_1":[], "c_diagnose_katalog_1":[],"mapped_katalog":[],"number_of_mappings":[]})

#connect to the PostgreSQL server
conn = None
params = config()
conn = psycopg2.connect(**params)
cur = conn.cursor()

for i in range(2,ICDCharite_Zeilenzahl):
    print(str(i)+"/"+str(ICDCharite_Zeilenzahl))
    current_icd_value=ICDCharite_worksheet[('A'+str(i))].value
    current_icd_catalog=ICDCharite_worksheet[('B'+str(i))].value
    #Depending whether the catalog is to be considered as well. Be aware, that catalg can be for example "ICD10 2007", which then can't be found in ATHENA vocabulary.
    #query = "".join(['SELECT'," * FROM public.concept WHERE vocabulary_id LIKE '",str(current_icd_catalog),"%' AND concept_code LIKE '", str(current_icd_value), "%'"])
    query = "".join(['SELECT'," * FROM public.concept WHERE vocabulary_id LIKE 'ICD%' AND concept_code LIKE '", str(current_icd_value), "%'"])
    cur.execute(query)
    QueryErgebnis = cur.fetchall()
    ICDKatalog =[]
    for z in range(0,cur.rowcount):
        ICDKatalog.append(QueryErgebnis[z][3])
    toappend = [current_icd_value, current_icd_catalog, str(ICDKatalog),cur.rowcount]
    ResultsDataframeICD = ResultsDataframeICD.append(pd.Series(toappend, index=ResultsDataframeICD.columns[:len(toappend)]), ignore_index=True)

ResultsDataframeICD.to_csv('ICD_full_results.csv')

# close the communication with the PostgreSQL
cur.close()
if conn is not None:
    conn.close()
