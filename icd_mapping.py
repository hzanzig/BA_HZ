#!/usr/bin/python
import psycopg2
import xlwt
from openpyxl import load_workbook
from xlwt import Workbook
from config import config

ICDCharite_object = load_workbook(r'C:\GitHub\BA_HZ\icd_diagnosis_full_charite.xlsx')
ICDCharite_result = Workbook()
sheet1 = ICDCharite_result.add_sheet('Sheet 1')
sheet2 = ICDCharite_result.add_sheet('Sheet 2')
sheet3 = ICDCharite_result.add_sheet('Sheet 3')
sheet4 = ICDCharite_result.add_sheet('Sheet 4') 
sheets = [sheet1,sheet2,sheet3,sheet4]
ICDCharite_worksheet = ICDCharite_object['icd_diagnosis_full_charite']
ICDCharite_Rowcount = ICDCharite_worksheet.max_row

#connection to database
conn = None
params = config()
# connect to the PostgreSQL server
print('Connecting to the PostgreSQL database...')
conn = psycopg2.connect(**params)
# create a cursor
cur = conn.cursor()
        
#set-up of the result sheet 
for i in range(0,4):
    sheets[i].write(0,0,"c_diagnose_1")
    sheets[i].write(0,1,"c_diagnose_catalog_1")
    sheets[i].write(0,2,"mapping_success")
    sheets[i].write(0,3,"mapped_catalog")
    sheets[i].write(0,4,"number_of_mappings")

#begin of the iteraton
for i in range(2,ICDCharite_Rowcount):
    print(str(i)+"/"+str(ICDCharite_Rowcount)) #print of iterations 
    icd_value=ICDCharite_worksheet[('A'+str(i))].value #getting the ICD code
    icd_catalog=ICDCharite_worksheet[('B'+str(i))].value #getting the ICD catalog ID 
    query = "".join(['SELECT'," * FROM public.concept WHERE vocabulary_id LIKE '",str(icd_catalog),"%' AND concept_code LIKE '", str(icd_value), "'"]) 
    cur.execute(query) #executing the SELECT query
    QueryResult = cur.fetchall()
    #Workbook can't extend the rowcount of 65,536. For number of ICD codes, several EXCEL sheets were therefore neccessary
    if i<65500: 
        if cur.rowcount == 0:
            sheet1.write(i-1, 0,  icd_value)
            sheet1.write(i-1, 1,  ICDCharite_worksheet[('B'+str(i))].value)
            sheet1.write(i-1, 2, 0)
        else:
            sheet1.write(i-1, 0,  icd_value)
            sheet1.write(i-1, 1,  ICDCharite_worksheet[('B'+str(i))].value)
            sheet1.write(i-1, 2, 1)            
            ICDCatalog =[]
            for z in range(0,cur.rowcount):
                ICDCatalog.append(QueryResult[z][3])
            sheet1.write(i-1, 3, str(ICDCatalog))
            sheet1.write(i-1,4,cur.rowcount)
    elif i<131000:
        if cur.rowcount == 0:
            sheet2.write(i-65499, 0,  icd_value)
            sheet2.write(i-65499, 1,  ICDCharite_worksheet[('B'+str(i))].value)
            sheet2.write(i-65499, 2, 0)
        else:
            sheet2.write(i-65499, 0,  icd_value)
            sheet2.write(i-65499, 1,  ICDCharite_worksheet[('B'+str(i))].value)
            sheet2.write(i-65499, 2, 1)
            ICDCatalog =[]
            for z in range(0,cur.rowcount):
                ICDCatalog.append(QueryResult[z][3])
            sheet2.write(i-65499, 3, str(ICDCatalog))
            sheet2.write(i-65499,4,cur.rowcount)
    elif i<196000:
        if cur.rowcount == 0:
            sheet3.write(i-130999, 0,  icd_value)
            sheet3.write(i-130999, 1,  ICDCharite_worksheet[('B'+str(i))].value)
            sheet3.write(i-130999, 2, 0)
        else:
            sheet3.write(i-130999, 0,  icd_value)
            sheet3.write(i-130999, 1,  ICDCharite_worksheet[('B'+str(i))].value)
            sheet3.write(i-130999, 2, 1)
            ICDCatalog =[]
            for z in range(0,cur.rowcount):
                ICDCatalog.append(QueryResult[z][3])
            sheet3.write(i-130999, 3, str(ICDCatalog))
            sheet3.write(i-130999,4,cur.rowcount)
    else:
        if cur.rowcount == 0:
            sheet4.write(i-195999, 0,  icd_value)
            sheet4.write(i-195999, 1,  ICDCharite_worksheet[('B'+str(i))].value)
            sheet4.write(i-195999, 2, 0)
        else:
            sheet4.write(i-195999, 0,  icd_value)
            sheet4.write(i-195999, 1,  ICDCharite_worksheet[('B'+str(i))].value)
            sheet4.write(i-195999, 2, 1)
            ICDCatalog =[]
            for z in range(0,cur.rowcount):
                ICDCatalog.append(QueryResult[z][3])
            sheet4.write(i-195999, 3, str(ICDCatalog))
            sheet4.write(i-195999,4,cur.rowcount)

ICDCharite_result.save('ICDCharite_result.xls') 

# close the communication with the PostgreSQL
cur.close()
if conn is not None:
    conn.close()
    print('Database connection closed.')