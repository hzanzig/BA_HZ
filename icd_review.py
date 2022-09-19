#!/usr/bin/python
from unitcurrent_excelsheet import skipIf
import psycopg2
import xlwt
from openpyxl import load_workbook
from xlwt import Workbook
from config import config

#loading of result excel file
ICDCharite_object = load_workbook(r'C:\GitHub\BA_HZ\ICDCharite_result.xlsx')
ICDCharite_result_worksheet1 = ICDCharite_object['Sheet 1']
ICDCharite_result_worksheet2 = ICDCharite_object['Sheet 2']
ICDCharite_result_worksheet3 = ICDCharite_object['Sheet 3']
ICDCharite_result_worksheet4 = ICDCharite_object['Sheet 4']
worksheets = [ICDCharite_result_worksheet1,ICDCharite_result_worksheet2,ICDCharite_result_worksheet3,ICDCharite_result_worksheet4]

#general matching and mismatching check
matches = 0
mismatches = 0
mismatchlist = []

#check whether the category matches as well
matchesCategories = 0
mismatchesCategories = 0
mismatchlistCategories = []

#going through all four result sheets
for i in range(0,4):
    for j in range(2,worksheets[i].max_row):
        current_excelsheet=worksheets[i]
        try:
            if current_excelsheet[('B'+str(j))].value == None:
                mismatchesCategories += 1       #if category in original file was empty, it's counted as a mismatch
                mismatchlistCategories.append(current_excelsheet[('A'+str(j))].value) 
            elif current_excelsheet[('D'+str(j))].value == None:    #if result category is emty, it's counted as mismatch
                mismatchesCategories += 1
                mismatchlistCategories.append(current_excelsheet[('A'+str(j))].value)
            elif ("ICD10" or "ICD-10") in (current_excelsheet[('B'+str(j))].value or current_excelsheet[('D'+str(j))].value):
                matchesCategories +=1
            elif ("ICD9" or "ICD-9") in (current_excelsheet[('B'+str(j))].value or current_excelsheet[('D'+str(j))].value):
                matchesCategories +=1    
        except:
            mismatchesCategories += 1
            mismatchlistCategories.append(current_excelsheet[('A'+str(j))].value)
            
        #checking for general matches and mismatches
        if current_excelsheet[('C'+str(j))].value == 1:
            matches +=1
        else:
            mismatches +=1
            mismatchlist.append(current_excelsheet[('A'+str(j))].value)

print("Matches: ", matches)
print("Mismatches: ", mismatches)
print(mismatchlist)

print("MatchesCategories: ", matchesCategories)
print("MismatchesCategories: ", mismatchesCategories)
print(mismatchlistCategories)

#comparing both mismatchlists to find the ICD codes which matched but with an invalid category
for i in range(0,len(mismatchlistCategories)):
    if mismatchlistCategories[i] in mismatchlist:
        skipIf
    else:
        print(mismatchlistCategories[i])


#The results of the review:

#Matches:  201918
#Mismatches:  108
#mismatchlist: ['E66.97', 'I31.88', 'K86.84', 'UUU', 'E66.27', 'E66.86', 'E66.88', 'G99.00', 'J98.10', 'K31.10', 'N13.63', 'N13.64', 'U69.54', '044.9', 'E66.98', 'G94.31', 'I77.80', 'J36.0', 'J98.12', 'K.07.2', 'K62.51', 'U69.55', 'I77.80', 'K65.09', 'N13.20', 'S70.86', '042.0', 'H.17.8', 'ICD10', 'INER ZWE', 'K72.18', 'N13.65', 'N99.10', '5-158.23', 'G94.32', 'J 35.8', 'J98.18', 'K31.18', 'K62.59', 'K70.48'', 'N13.21', 'U69.51', ' ', ' ', '042.1', '042.2', 'E66.06', 'G94.30', 'G94.39', 'K31.11', 'K31.12', 'K65.09', 'M9391/3', 'N13.61', 'N99.18', 'U69.6', 'UUU', 'V30.-', 'E66.87', 'N13.62', 'N13.66', 'T08.X0', 'U69.50', '043.2', '540-543', 'E66.08', 'E66.28', 'H35.39', 'K65.00', 'N13.68', 'S', ' ', ' ', '482.-', 'CA', 'E66.96', 'K86.80']

#MatchesCategories: 188477
#MismatchesCategories: 112
#mismatchlistCategories: ['E66.97', 'I31.88', 'J35.1', 'K86.84', 'UUU', 'E66.27', 'E66.86', 'E66.88', 'G99.00', 'J98.10', 'K31.10', 'N13.63', 'N13.64', 'R13', 'U69.54', '044.9', 'E66.98', 'G94.31', 'I77.80', 'J36.0', 'J98.12', 'K.07.2', 'K62.51', 'U69.55', 'I77.80', 'K65.09', 'N13.20', 'S70.86', '042.0', 'H.17.8', 'ICD10', 'INER ZWE', 'K72.18', 'N13.65', 'N99.10', '5-158.23', 'G94.32', 'J 35.8', 'J98.18', 'K31.18', 'K62.59', 'K70.48', 'N13.29', 'N13.60', 'STRUKTIO', '042.9', 'E 950', 'I77.88', 'K65.00', 'U69.53', 'UUU', 'H33-0', 'J 35.0', 'K83.08', 'K86.82', 'K86.88', 'R35.2', 'T08.X0', 'U13.2', '043.3', 'E66.07', 'E66.26', 'H35.39', 'I77.88', 'K83.00', 'NKS.8', 'U69.52', '043.1', 'D18.051', 'I31.80', 'K62.50', 'K70.42', 'K86.83', 'N13.21', 'U69.51', ' ', ' ', '042.1', '042.2', 'E66.06', 'G94.30', 'G94.39', 'K31.11', 'K31.12', 'K65.09', 'M9391/3', 'N13.61', 'N99.18', 'T81.0', 'U69.6', 'UUU', 'V30.-', 'E66.87', 'N13.62', 'N13.66', 'T08.X0', 'U69.50', '043.2', '540-543', 'E66.08', 'E66.28', 'H35.39', 'K65.00', 'N13.68', 'S', ' ', ' ', '482.-', 'CA', 'E66.96', 'K86.80']

#J35.1
#R13
#T81.0
#H26